#!/usr/bin/env python
# coding: utf-8

# In[1]:


import os, shutil, re, csv, openpyxl, math
import pandas as pd
import numpy as np
from openpyxl.styles import PatternFill


# In[2]:


def get_color_row_idx(file):
    workbook = openpyxl.load_workbook(file)
    worksheet = workbook['Individual']
    start_col_idx = 0
    total_col = worksheet.max_column # 48columns need to be extracted
    color_map = openpyxl.styles.colors.COLOR_INDEX
    
    for cell in worksheet['A']:
        if cell.value == 'Time(min)':
            start_row_idx = cell.row
            print("Time(min) start:",start_row_idx) # group1.xlsm shoud be : 50
    for column1 in worksheet.iter_cols(min_col=start_col_idx, max_col = start_col_idx+1, min_row =start_row_idx+1 ):
        for cell in column1:
            color_index = cell.fill.start_color.index
            rgb = color_map[int(color_index)]
            if rgb == '00FF0000':
                end_time_row_idx = cell.row
                print("Index of 1st red BG color:",end_time_row_idx)
                break
    workbook.close()
    return start_row_idx ,end_time_row_idx


# In[3]:


def dis2mm(dis_value):
    return round(11/130*dis_value,3)


# In[4]:


def get_groupID(file):
    workbook = openpyxl.load_workbook(file)
    worksheet = workbook['Setting']
    if worksheet.cell(row=15, column=18).value == 1:
        return 1
    else:
        return 2
    
def colorBYgroup():
    odd_col_g1 = []
    even_col_g1 = []
    yellow_fill = PatternFill(start_color="9BC2E5", end_color="9BC2E5", fill_type="solid")
    for i in range (2,99,2):
        odd_col_g1.append(i+1)
        even_col_g1.append(i)  
    del even_col_g1[24]
    
    for wb in ['Dt_bf_5PM.xlsx','DT_10halfto1PM.xlsx']:
        workbook = openpyxl.load_workbook(wb)
        for worksheet in ['Individual','Distance']:
            sheet = workbook[worksheet]

            if colorby[0] == 1:
                print('first file well1 = G1')
                for row in sheet.iter_rows():
                    for column_number, cell in enumerate(row, start=1):
                        if column_number in even_col_g1 and cell.value is not None:
                            # 如果单元格不为空，设置背景颜色为黄色
                            cell.fill = yellow_fill

            elif colorby[0] == 2:
                print('first file well1 = G2')
                for row in sheet.iter_rows():
                    for column_number, cell in enumerate(row, start=1):
                        if column_number in even_col_g1 and cell.value is not None:
                            # 如果单元格不为空，设置背景颜色为黄色
                            cell.fill = yellow_fill    

        workbook.save(wb)
        workbook.close()    
    return 'coloring column DONE!!'


# ## step1： 
# ## get before red data from individual and distance
# ## transform distance pixel value to mm value

# In[5]:


fullpath = os.getcwd()
fileall = os.listdir(fullpath)
if 'Dt_bf_5PM.xlsx' in fileall:
    os.remove('Dt_bf_5PM.xlsx')
    os.remove('DT_10halfto1PM.xlsx')
    print('remove previous groupfile.xlsx')
excelfile = []
individual_nored_data = pd.DataFrame()
distance_df = pd.DataFrame()
colorby = []
    
for File in os.listdir(fullpath):
        if File.endswith('.xlsm'):
            #print(File)
            excelfile.append(File)


# In[6]:


start_row_idx = []
end_row_idx = []            
for file in excelfile:       
    print("data extracting from : ",file)

    # openpyxl read colored cell, return the first colored time index
    print('Individual data processing ---')
    start_idx ,end_idx = get_color_row_idx(file)
    start_row_idx.append(start_idx)
    end_row_idx.append(end_idx)
    
    ## read file
    individual = pd.read_excel(fullpath+'/'+file, sheet_name = 'Individual')
    distance = pd.read_excel(fullpath+'/'+file, sheet_name = 'Distance')

    individual_nored_data = pd.concat([individual_nored_data,individual.iloc[start_idx-2:end_idx-2,:]],axis = 1)
    distance_val = distance.iloc[start_idx-2:end_idx-2,:]
    
    # set mm value round to two decimal places
    
    distance_val.columns = distance.iloc[48]
    distance_val = distance_val.iloc[1:].reset_index(drop=True)
    distance_val = distance_val.astype(float)
    distance_val.iloc[:, 1:] = distance_val.iloc[:, 1:] *11/130
    distance_val.iloc[:, 1:] = distance_val.iloc[:, 1:].round(2)
    distance_df = pd.concat([distance_df,distance_val],axis = 1)
    
    # decide the grouping sequence of file 
    colorby.append(   get_groupID(file)   )
    
with pd.ExcelWriter(fullpath +'/Dt_bf_5PM.xlsx', engine = 'xlsxwriter') as writer:
    
    individual_nored_data.to_excel(writer ,sheet_name = 'Individual',index = False ,header = False)
    distance_df.to_excel(writer ,sheet_name = 'Distance',index = False ,header = True)
    
    worksheet = writer.sheets['Individual']
    worksheet.set_column("A:Z", 10)


# In[ ]:





# In[7]:


individual_nored_data = pd.DataFrame()
distance_df = pd.DataFrame()
loop = 0
for file in excelfile:
    print(file)
    individual = pd.read_excel(fullpath+'/'+file, sheet_name = 'Individual' )
    distance = pd.read_excel(fullpath+'/'+file, sheet_name = 'Distance' )
    print(start_row_idx[loop])
    header = individual.iloc[start_row_idx[loop]-2,:]
    
    start_10half_idx = end_row_idx[loop] +300-2
    end_1pm_idx = start_10half_idx+150
    
    loop += 1
    
    individual_10to1 = individual.iloc[start_10half_idx:end_1pm_idx, :]
    distance_10to1 = distance.iloc[start_10half_idx:end_1pm_idx,:]
    individual_10to1.columns = header
    distance_10to1.columns = header
    individual_10to1 = individual_10to1.reset_index(drop= True)
    distance_10to1 = distance_10to1.reset_index(drop= True)
    
    print(individual_10to1.head(2))
    
    individual_nored_data = pd.concat([individual_nored_data,individual_10to1],axis = 1)
    #distance_val = distance.iloc[start_row_idx-2:end_row_idx-2,:]
    
    # set mm value round to two decimal places
    #distance_10to1 = distance_10to1.iloc[1:].reset_index(drop=True)
    distance_10to1 = distance_10to1.astype(float)
    distance_10to1.iloc[:, 1:] = distance_10to1.iloc[:, 1:] *11/130
    distance_10to1.iloc[:, 1:] = distance_10to1.iloc[:, 1:].round(2)
    distance_df = pd.concat([distance_df,distance_10to1],axis = 1)
    
    # decide the grouping sequence of file 
    colorby.append(   get_groupID(file)   )
    
with pd.ExcelWriter(fullpath +'/DT_10halfto1PM.xlsx', engine = 'xlsxwriter') as writer:
    print('writing data to new excel')
    individual_nored_data.to_excel(writer ,sheet_name = 'Individual',index = False ,header = True)
    distance_df.to_excel(writer ,sheet_name = 'Distance',index = False ,header = True)
    
    worksheet = writer.sheets['Individual']
    worksheet.set_column("A:CT", 10)
    worksheet = writer.sheets['Distance']
    worksheet.set_column("A:CT", 10)    
    


# In[ ]:





# ## Step 2 color columns by group ID, Group 1 are colored with blue background

# In[8]:


colorBYgroup()


# ## STEP 3: calculate each 30 mins data by fish_xx
# 

# Data before 5PM 

# In[9]:


def sum_data_grouping(A,B,FLAG=1):
    WT_col = []
    A5_col= []
    
    if FLAG ==1:
        for i in range(min(len(A), len(B))):
            if i % 2 == 0:  
                WT_col.append(A[i])
                A5_col.append(B[i])
            else:
                WT_col.append(B[i])
                A5_col.append(A[i])
    elif FLAG ==2:
        for i in range(min(len(A), len(B))):
            if i % 2 == 0:  # 如果索引是偶数
                WT_col.append(B[i])
                A5_col.append(A[i])
            else:
                WT_col.append(A[i])
                A5_col.append(B[i])        
    return WT_col , A5_col


# In[10]:


for worksheet in ['Individual','Distance']:
    print('working on sheet:%s'%worksheet)
    every30_df = []
    cal_every30 = pd.DataFrame()
    cal_total = pd.DataFrame()
    groupA = [] # left group
    groupB = [] # right group
    totalA = []
    totalB = []
    df = pd.read_excel('Dt_bf_5PM.xlsx',sheet_name = worksheet)

    ## calculate total by minimun row
    nan_row_idx = np.where(np.isnan(df))[0][0] # return the first row where NAN appear
    nan_col_idx = np.where(np.isnan(df))[1][1]
    print(nan_row_idx,nan_col_idx )  # 106
    df_min = df.iloc[:nan_row_idx]
    totalA = df_min.iloc[:,1:49].sum().round(2).tolist()
    totalB = df_min.iloc[:,50:].sum().round(2).tolist()    
    
    if colorby[0] == 1:
        WT_total , A5_total = sum_data_grouping(totalA, totalB, 1)

    elif colorby[0] == 2:
        WT_total , A5_total = sum_data_grouping(totalA, totalB, 2)
    
    cal_total = pd.DataFrame({'WT_total':totalA,'A5_total':totalB})
    cal_total = cal_total.set_index(df.columns[1:49])    
    
    for i in range (math.ceil(df.shape[0]/30)):
        every30_df.append( (df.iloc[i*30:,:] if i == math.ceil(df.shape[0]/30) -1 else df.iloc[i*30:(i+1)*30,:] )  )
        if nan_row_idx >i*30 and nan_row_idx <(i+1)*30:
            print('this is the %s part of interval column'%(i+1))
            colored_flg = i+1  # record the column that need to be colored
            
            
    for index, sub_df in enumerate(every30_df):
        groupA = sub_df.iloc[:,1:49].sum().tolist()
        groupB = sub_df.iloc[:,50:].sum().round(2).tolist()# correct
        # group1 =WT, group2 = A5
        if colorby[0] == 1:
            WT_col , A5_col = sum_data_grouping(groupA, groupB, 1)
            
        elif colorby[0] == 2:
            WT_col , A5_col = sum_data_grouping(groupA, groupB, 2)
            
            
        cal_30 = pd.DataFrame({'WT_%sto%s'%(index*30,(index+1)*30):WT_col,'A5_%sto%s'%(index*30,(index+1)*30):A5_col})
        cal_30 = cal_30.set_index(df.columns[1:49])
        cal_every30 = pd.concat([cal_every30,cal_30],axis = 1)
    
    cal_total = pd.concat([cal_every30,cal_total],axis = 1)
    
    # calculate mean and SEM for the total dataframe
    column_means = cal_total.mean()
    sem_result = cal_total.sem()
    cal_total.loc['Group Mean'] = column_means
    cal_total.loc['Group SEM'] = sem_result
    
    with pd.ExcelWriter('Dt_bf_5PM.xlsx', mode='a', engine='openpyxl') as writer:
        cal_total.to_excel(writer, sheet_name='%s_SUM30min'%worksheet, index=True)
        print('new sheet writing')
print('DONE!!')


# In[11]:


## color the mini-interval data


# In[12]:


workbook = openpyxl.load_workbook('Dt_bf_5PM.xlsx')
yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
# 选择要操作的工作表（sheet）
for wb in ['Individual_SUM30min','Distance_SUM30min']:
    sheet = workbook[wb]

    # 遍历第八列和第九列的单元格，根据条件填充颜色
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row-2, min_col=colored_flg*2, max_col=colored_flg*2+1):
        for cell in row:
            if colorby[0]==1:
                if cell.row %2 ==0 and cell.column == colored_flg*2:
                    cell.fill = yellow_fill
                elif cell.row %2 !=0 and cell.column == colored_flg*2+1:
                    cell.fill = yellow_fill
                    
            elif colorby[0]==2:
                if cell.row %2 ==0 and cell.column == colored_flg*2+1:
                    cell.fill = yellow_fill

                elif cell.row %2 !=0 and cell.column == colored_flg*2:
                    cell.fill = yellow_fill
                  


workbook.save('Dt_bf_5PM.xlsx')


# Data between 10:30PM to 1PM

# In[13]:


for worksheet in ['Individual','Distance']:
    print('working on sheet:%s'%worksheet)
    every30_df = []
    cal_every30 = pd.DataFrame()
    cal_total = pd.DataFrame()
    groupA = [] # left group
    groupB = [] # right group
    totalA = []
    totalB = []
    df = pd.read_excel('DT_10halfto1PM.xlsx',sheet_name = worksheet)

    
    ## calculate total by minimun row

    totalA = df.iloc[:,1:49].sum().round(2).tolist()
    totalB = df.iloc[:,50:].sum().round(2).tolist()    
    
    if colorby[0] == 1:
        WT_total , A5_total = sum_data_grouping(totalA, totalB, 1)
    elif colorby[0] == 2:
        WT_total , A5_total = sum_data_grouping(totalA, totalB, 2)
    
    cal_total = pd.DataFrame({'WT_total':WT_total,'A5_total':A5_total})
    cal_total = cal_total.set_index(df.columns[1:49])    
    
    for i in range (math.ceil(df.shape[0]/30)):
        every30_df.append( (df.iloc[i*30:,:] if i == math.ceil(df.shape[0]/30) -1 else df.iloc[i*30:(i+1)*30,:] )  )
    
    for index, sub_df in enumerate(every30_df):
        groupA = sub_df.iloc[:,1:49].sum().tolist()
        groupB = sub_df.iloc[:,50:].sum().round(2).tolist()# correct

        # group1 =WT, group2 = A5
        if colorby[0] == 1:
            WT_col , A5_col = sum_data_grouping(groupA, groupB, 1)
            
        elif colorby[0] == 2:
            WT_col , A5_col = sum_data_grouping(groupA, groupB, 2)
                    
        cal_30 = pd.DataFrame({'WT_%sto%s'%(index*30,(index+1)*30): WT_col ,'A5_%sto%s'%(index*30,(index+1)*30): A5_col})  
        cal_30 = cal_30.set_index(df.columns[1:49])
        cal_every30 = pd.concat([cal_every30,cal_30],axis = 1)
    
    cal_total = pd.concat([cal_every30,cal_total],axis = 1)
    
    # calculate mean and SEM for the total dataframe
    column_means = cal_total.mean()
    sem_result = cal_total.sem()
    cal_total.loc['Group Mean'] = column_means
    cal_total.loc['Group SEM'] = sem_result
    
    with pd.ExcelWriter('DT_10halfto1PM.xlsx', mode='a', engine='openpyxl') as writer:
        cal_total.to_excel(writer, sheet_name='%s_SUM30min'%worksheet, index=True)
        
print('DONE!!')


# In[ ]:





# In[ ]:





# In[ ]:





# In[ ]:





# In[ ]:





# In[ ]:





# In[ ]:




